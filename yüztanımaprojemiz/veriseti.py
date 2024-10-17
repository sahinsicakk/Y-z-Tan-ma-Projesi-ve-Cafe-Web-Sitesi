from tkinter import * #Arayüz kütüphanesi
from tkinter import messagebox
import cv2 # Görüntü işleme kütüphanesi
from openpyxl import * # Excelden veri eklemek için
import sqlite3 # Veritabanı kütüphanesi
ana_pencere =Tk() # Ara yüz oluşturma fonksiyonu
ana_pencere.title("Yoklama Sistemi") # Arayüze isim verme
ana_pencere.configure(background='turquoise') # Arayüze renk ekleme
yazi_tipi = "Helvetica 18 bold" # Arayüz yazı tipi
ana_pencere.geometry("700x400") # Arayüz boyutu belirliyoruz

newUserPhoto = PhotoImage(file=r"new_user.png")# buton resimleri ekleniyor.
yeniOgrenci = newUserPhoto.subsample(2, 2) # Resmi yeniden boyutlandırma
yoklama = PhotoImage(file=r"new_user.png")# buton resimleri ekleniyor.
yoklamaResim = yoklama.subsample(2, 2)# Resmi yeniden boyutlandırma
baslik = Label(ana_pencere, text="Yoklama Sistemine Hoşgeldiniz", font=yazi_tipi)# arayüz üzerine başlığımızı yazıyoruz
baslik.grid(column=0, row=1) # Başlığın konumunu belirliyoruz
baglanti = sqlite3.connect("personeller.db") # Veritabanımızı kuruyoruz
im = baglanti.cursor() # Veritabanı bağlantımızı im değişkenine atayıp istediğimiz yerde kullanıyoruz
im.execute("CREATE TABLE IF NOT EXISTS personeller(ADI TEXT, SOYADI TEXT,GMAİL TEXT,  PAROLA TEXT)")# Tablomuzu oluşturduk
baglanti.commit()# Veritabanına bağlandık
def kaydet_bigiler(ADI, SOYADI, GMAİL, PAROLA):
    im.execute("INSERT INTO personeller(ADI, SOYADI, GMAİL , PAROLA ) VALUES (?,?,?,?)", (ADI, SOYADI, GMAİL, PAROLA)) #Veritabanına bilgileri ekliyoruz
    baglanti.commit() # Veritabanına bağlandık
def kaydet():
    kayıt = Tk() # Kaydetmek için yeni bir arayüz ekledim
    kayıt.configure(background='light green') # Arayüze renk ekleme
    kayıt.geometry("500x300") # Arayüz boyutu belirliyoruz
    yazi_tipi = "Helvetica 18 bold"  # Arayüz yazı tipi
    kayıt.title("kayıt formu") # Arayüze isim verme
    kayit_baslik = Label(kayıt, text="Yeni ŞEF KAYIT ET", bg="light green", font=yazi_tipi)# arayüz üzerine başlığımızı yazıyoruz
    kayit_baslik.grid(row=0, column=1)  #Başlığın konumunu belirliyoruz
    adi = Label(kayıt, text="ADI", bg="light green") # Arayüze istenen bilgileri ekliyoruz
    adi.grid(row=1, column=0)# Metinin konumunu belirliyoruz
    gir1 = Entry(kayıt) # Metin kutusunu açar
    gir1.grid(row=1, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz
    soadi = Label(kayıt, text="SOYADI", bg="light green")# Arayüze istenen bilgileri ekliyoruz
    soadi.grid(row=2, column=0)##Metinin konumunu belirliyoruz
    gir2 = Entry(kayıt)# Metin kutusunu açar
    gir2.grid(row=2, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz
    gmail = Label(kayıt, text="MAİL", bg="light green")# Arayüze istenen bilgileri ekliyoruz
    gmail.grid(row=3, column=0)#Metinin konumunu belirliyoruz
    gir3 = Entry(kayıt)# Metin kutusunu açar
    gir3.grid(row=3, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz
    parola = Label(kayıt, text="PAROLA", fg="red", bg="light green")# Arayüze istenen bilgileri ekliyoruz
    parola.grid(row=4, column=0)#Metinin konumunu belirliyoruz
    gir4 = Entry(kayıt)# Metin kutusunu açar
    gir4.grid(row=4, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz

    def degerle():
        kaydet_bigiler(gir1.get(), gir2.get(), gir3.get(), gir4.get())#Arayüzde girdiğimiz bilgileri kaydetbilgileri fonksiyonuna gönderiyoruz
        kod = 0 # bir değişken tanidık
        adi1= gir1.get() # gir1 deki bilgiyi bana getir
        im.execute("SELECT COUNT(*) FROM personeller")# Veritabanındaki bilgileri seçiyoruz
        baglanti.commit()# Veritabanına bağlandık
        kontrol = im.fetchall() #veritabanı içindeki tek satır bilgilerini getirir
        for i in kontrol:
            ad_kontrol = i[0] # İ nin 0 indexsine bakıyor
            if ad_kontrol == adi1:
                kod = 1
        if kod == 0:
            messagebox.showinfo("Kayıt Tamamlandı!", "Şef Kayıt Edildi!")

        kayıt.destroy()

    kaydet_btn = Button(kayıt, text="Kaydet", fg="White", bg="Red", command=degerle)#Buttonumuzu ekliyoruz
    kaydet_btn.grid(row=8, column=0, ipadx="50", ipady="20")# Buttonun konumunu belirliyoruz

    kayıt.mainloop() # Ve arayüzümüzü kapatıyoruz



def YeniOgrenciKaydet():
    kayit_penceresi=Tk()
    kayit_penceresi.configure(background='light green')# arkaplan rengi
    kayit_penceresi.title("Kayıt Formu")# pencere başlığı
    kayit_penceresi.geometry("500x300")# pencere boyutu
    kayit_baslik = Label(kayit_penceresi, text="Yeni Personel Kayıt Et", bg="light green", font=yazi_tipi)# arayüz üzerine başlığımızı yazıyoruz
    # etiketler olusturuluyor.
    ogrenci_no = Label(kayit_penceresi, text="Personel No", bg="light green")# Arayüze istenen bilgileri ekliyoruz
    ad = Label(kayit_penceresi, text="Ad", bg="light green")# Arayüze istenen bilgileri ekliyoruz
    soyad = Label(kayit_penceresi, text="Soyad", bg="light green")# Arayüze istenen bilgileri ekliyoruz
    # yerlesşimler yapılıyor.
    kayit_baslik.grid(row=0, column=1)#Metinin konumunu belirliyoruz
    ogrenci_no.grid(row=1, column=0)#Metinin konumunu belirliyoruz
    ad.grid(row=2, column=0)#Metinin konumunu belirliyoruz
    soyad.grid(row=3, column=0)#Metinin konumunu belirliyoruz
    # giriş yapılacak alanlar ekleniyor.
    ogrenci_no_field = Entry(kayit_penceresi)# Metin kutusunu açar
    ad_field = Entry(kayit_penceresi)# Metin kutusunu açar
    soyad_field = Entry(kayit_penceresi)# Metin kutusunu açar
    ogrenci_no_field.grid(row=1, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz
    ad_field.grid(row=2, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz
    soyad_field.grid(row=3, column=1, ipadx="100")#Metinin kutusunun konumunu belirliyoruz
    wb = load_workbook('personel_listesi.xlsx') #excel dosyasını oluşturuyoruz

    sheet = wb.active# excele kaydetmek için bir çalışma sayfası ayrılıyor.
    def YuzAlma():
        try:
            cam = cv2.VideoCapture(0,cv2.CAP_DSHOW)# Kameramızı açıyoruz
            detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')# cv2 nin hazır yüz tanıma dosyası okuyoruz
            Id = ogrenci_no_field.get() # metin kutusuna yazılan bilgiyi bana getir
            sampleNum = 0 #resim sayısını belirlemek için bir değişken tanıdım
            while (True):
                ret, img = cam.read() #ret ve img diye iki değişken tanıttım ve camerada okumasını sağladım
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) #img değişkenine gri rengi veriyoruz
                faces = detector.detectMultiScale(gray, 1.3, 5)# yüzü buluyoruz
                for (x, y, w, h) in faces:
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2) #yüzü kare içine alıyoruz ve karenin resmi belirliyoruz
                    sampleNum = sampleNum + 1 # Yukardaki deişkenimizi birer artırıyoruz
                    cv2.imwrite("datasets/" + Id + '.' + str(sampleNum) + ".jpg", gray[y:y + h, x:x + w])# Çekilen resimleri jpg formatında datasets klasorüne kaydediyoruz
                    cv2.imshow('YUZ TARAMA', img)# Yüzleri çekmek için kameramzı açıyoruz

                if cv2.waitKey(100) & 0xFF == ord('q'): #  q bastığı zaman çıkmasını sağlıyoruz
                    break

                elif sampleNum > 50: #resim sayısı 50 olduğu zaman program dursub
                    break

            cam.release()
            cv2.destroyAllWindows()# Bütün penceler kapanıyor
        except:
            print("Hata")

    def clear():
        # metin girişleri yapılacak alanları temizleyen metot.
        ad_field.delete(0, END)
        soyad_field.delete(0, END)
        ogrenci_no_field.delete(0, END)

    def OgrenciyiKaydet():

        if ad_field.get() == "" or soyad_field.get() == "" or ogrenci_no_field.get() == "":
            print("Tüm alanlar doldurulmadı!")
            messagebox.showerror("Hata!", "Lütfen tüm alanları doldurun!")
        else:
            try:
                current_row = sheet.max_row
                sheet.cell(row=current_row + 1, column=1).value = ogrenci_no_field.get()
                sheet.cell(row=current_row + 1, column=2).value = ad_field.get()
                sheet.cell(row=current_row + 1, column=3).value = soyad_field.get()
                wb.save('personel_listesi.xlsx')
                ad_field.focus_set()

            except:
                messagebox.showerror("Hata", "Kayıt Başarısız!")
            else:
                # kayıt gerçekleşirse bir bilgi mesajı çıkıyor ve clear metodu çağırılarak alanlar temizleniyor.
                messagebox.showinfo("Kayıt Tamamlandı!", "Personel Kayıt Edildi!")
                clear()

    kaydet_btn = Button(kayit_penceresi, text="Kaydet", fg="White", bg="Red", command=OgrenciyiKaydet)
    kaydet_btn.grid(row=8, column=0, ipadx="50", ipady="20")
    yuz_tara_btn = Button(kayit_penceresi, text="Yüz Tara", command=YuzAlma)
    yuz_tara_btn.grid(row=8, column=1, ipadx="50", ipady="20")
    kayit_penceresi.mainloop()


yeni_ogrenci_btn = Button(ana_pencere, text="Personel kaydet", font=yazi_tipi, image=yeniOgrenci, compound=TOP,
                          bg="light green",
                          fg="blue", command=YeniOgrenciKaydet)
yeni_ogrenci_btn.grid(column=0, row=5)
yoklamaAl_btn = Button(ana_pencere, text="Şef kaydet", font=yazi_tipi, image=yoklamaResim, compound=TOP,
                       bg="light green", fg="blue", command=kaydet)
yoklamaAl_btn.grid(column=1, row=5)

ana_pencere.mainloop()